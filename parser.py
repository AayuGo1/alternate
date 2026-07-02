"""
parser.py
Converts an openpyxl Workbook (Monthly KPI Summary Sheet_April_GNSC.xlsx)
into a single long-format pandas DataFrame.

Sheet layout is detected structurally (by locating the header row that
contains month names / dates) rather than hardcoded, since the number of
hierarchy columns (Category / Subcategory / Metric) differs between sheets
in this workbook (e.g. "H&S" uses Category + Metric, "Environment" uses
Category + Subcategory + Metric). No KPI names, categories, or values are
hardcoded or fabricated.

This module contains NO Streamlit code, NO KPI calculations, and NO charts.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Final

import pandas as pd
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger: Final[logging.Logger] = logging.getLogger(__name__)

_MONTH_ABBR_TO_NUM: Final[dict[str, int]] = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_NUM_TO_MONTH_ABBR: Final[dict[int, str]] = {
    v: k.capitalize() for k, v in _MONTH_ABBR_TO_NUM.items()
}

_TARGET_HEADER_TOKENS: Final[set[str]] = {"target", "bu", "budget"}
_MIN_MONTH_HITS: Final[int] = 3
_HEADER_SEARCH_ROWS: Final[int] = 6
_UNIT_PATTERN: Final[re.Pattern[str]] = re.compile(r"\[([^\[\]]+)\]\s*$")

LONG_DF_COLUMNS: Final[list[str]] = [
    "Sheet", "Category", "Subcategory", "Metric", "Month",
    "Value", "Unit", "Target", "YTD", "MTD", "Year",
]


@dataclass
class _SheetLayout:
    """Structural layout of a single worksheet, detected (not hardcoded)."""

    header_row: int
    category_col: int | None
    subcategory_col: int | None
    metric_col: int
    month_cols: list[int]
    ytd_col: int | None
    target_col: int | None
    mtd_col: int | None
    year_map: dict[int, int | None] = field(default_factory=dict)


def _is_month_like(value: Any) -> bool:
    """Return True if a header cell value represents a month (date or name)."""
    if isinstance(value, (date, datetime)):
        return True
    if isinstance(value, str):
        token = value.strip()[:3].lower()
        return token in _MONTH_ABBR_TO_NUM
    return False


def _month_label_and_num(value: Any) -> tuple[str, int]:
    """Return (3-letter month abbreviation, month number) for a header cell."""
    if isinstance(value, (date, datetime)):
        return value.strftime("%b"), value.month
    token = str(value).strip()[:3].lower()
    month_num = _MONTH_ABBR_TO_NUM[token]
    return _NUM_TO_MONTH_ABBR[month_num], month_num


def _build_merge_lookup(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    Map every cell coordinate inside a merged range to the value held in
    that range's top-left cell, enabling forward-fill of merged category /
    subcategory cells exactly as stored in the workbook.
    """
    lookup: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = top_left_value
    return lookup


def _detect_header_row(ws: Worksheet) -> int | None:
    """Find the row containing month names/dates (the KPI table header row)."""
    best_row: int | None = None
    best_hits = 0
    max_row_to_scan = min(_HEADER_SEARCH_ROWS, ws.max_row or 0)
    for row in range(1, max_row_to_scan + 1):
        hits = sum(
            1
            for col in range(2, (ws.max_column or 0) + 1)
            if _is_month_like(ws.cell(row=row, column=col).value)
        )
        if hits > best_hits:
            best_hits = hits
            best_row = row
    if best_hits >= _MIN_MONTH_HITS:
        return best_row
    return None


def _detect_layout(ws: Worksheet) -> _SheetLayout | None:
    """Detect column roles (Category/Subcategory/Metric/months/YTD/Target/MTD)."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        return None

    max_col = ws.max_column or 0

    first_month_col: int | None = None
    for col in range(2, max_col + 1):
        if _is_month_like(ws.cell(row=header_row, column=col).value):
            first_month_col = col
            break
    if first_month_col is None:
        return None

    month_cols: list[int] = []
    col = first_month_col
    while col <= max_col and _is_month_like(ws.cell(row=header_row, column=col).value):
        month_cols.append(col)
        col += 1

    ytd_col: int | None = None
    target_col: int | None = None
    mtd_col: int | None = None
    for scan_col in range(col, max_col + 1):
        header_value = ws.cell(row=header_row, column=scan_col).value
        if not isinstance(header_value, str):
            continue
        token = header_value.strip().lower()
        if token == "ytd" and ytd_col is None:
            ytd_col = scan_col
        elif token in _TARGET_HEADER_TOKENS and target_col is None:
            target_col = scan_col
        elif token == "mtd" and mtd_col is None:
            mtd_col = scan_col

    label_cols = list(range(2, first_month_col))
    if not label_cols:
        return None

    metric_col = label_cols[-1]
    hierarchy_cols = label_cols[:-1]
    category_col = hierarchy_cols[0] if len(hierarchy_cols) >= 1 else None
    subcategory_col = hierarchy_cols[1] if len(hierarchy_cols) >= 2 else None

    year_map = _build_year_map(ws, header_row, month_cols)

    return _SheetLayout(
        header_row=header_row,
        category_col=category_col,
        subcategory_col=subcategory_col,
        metric_col=metric_col,
        month_cols=month_cols,
        ytd_col=ytd_col,
        target_col=target_col,
        mtd_col=mtd_col,
        year_map=year_map,
    )


def _build_year_map(ws: Worksheet, header_row: int, month_cols: list[int]) -> dict[int, int | None]:
    """
    Resolve the calendar year for each month column using the row directly
    above the header row (which holds year labels, forward-filled left to
    right), falling back to the header cell's own year if it is a date.
    """
    year_row = header_row - 1
    year_map: dict[int, int | None] = {}
    current_year: int | None = None
    for col in month_cols:
        raw_year = ws.cell(row=year_row, column=col).value if year_row >= 1 else None
        if raw_year is not None:
            try:
                current_year = int(raw_year)
            except (TypeError, ValueError):
                pass
        header_value = ws.cell(row=header_row, column=col).value
        if isinstance(header_value, (date, datetime)):
            year_map[col] = header_value.year
        else:
            year_map[col] = current_year
    return year_map


def _split_metric_and_unit(raw_metric: str) -> tuple[str, str | None]:
    """Extract a trailing bracketed unit from a metric name, if present."""
    match = _UNIT_PATTERN.search(raw_metric)
    if not match:
        return raw_metric.strip(), None
    unit = match.group(1).strip()
    metric_name = _UNIT_PATTERN.sub("", raw_metric).strip()
    return metric_name, unit


def _get_merged_or_direct(
    lookup: dict[tuple[int, int], Any], ws: Worksheet, row: int, col: int
) -> Any:
    """Return a cell's value, resolving merged-cell membership first."""
    if (row, col) in lookup:
        return lookup[(row, col)]
    return ws.cell(row=row, column=col).value


def _parse_sheet(ws: Worksheet, layout: _SheetLayout) -> list[dict[str, Any]]:
    """Extract long-format records from a single worksheet using its layout."""
    merge_lookup = _build_merge_lookup(ws)
    records: list[dict[str, Any]] = []

    current_category: Any = None
    current_subcategory: Any = None

    for row in range(layout.header_row + 1, (ws.max_row or 0) + 1):
        raw_metric = _get_merged_or_direct(merge_lookup, ws, row, layout.metric_col)
        if raw_metric is None or str(raw_metric).strip() == "":
            continue

        if layout.category_col is not None:
            cat_value = _get_merged_or_direct(merge_lookup, ws, row, layout.category_col)
            if cat_value is not None:
                current_category = cat_value

        if layout.subcategory_col is not None:
            subcat_value = _get_merged_or_direct(merge_lookup, ws, row, layout.subcategory_col)
            if subcat_value is not None:
                current_subcategory = subcat_value

        metric_name, unit = _split_metric_and_unit(str(raw_metric))

        target_value = (
            ws.cell(row=row, column=layout.target_col).value
            if layout.target_col is not None
            else None
        )
        ytd_value = (
            ws.cell(row=row, column=layout.ytd_col).value
            if layout.ytd_col is not None
            else None
        )
        mtd_value = (
            ws.cell(row=row, column=layout.mtd_col).value
            if layout.mtd_col is not None
            else None
        )

        for month_col in layout.month_cols:
            header_value = ws.cell(row=layout.header_row, column=month_col).value
            month_abbr, _ = _month_label_and_num(header_value)
            year = layout.year_map.get(month_col)
            month_label = f"{month_abbr}-{year}" if year is not None else month_abbr

            records.append(
                {
                    "Sheet": ws.title,
                    "Category": current_category,
                    "Subcategory": current_subcategory,
                    "Metric": metric_name,
                    "Month": month_label,
                    "Value": ws.cell(row=row, column=month_col).value,
                    "Unit": unit,
                    "Target": target_value,
                    "YTD": ytd_value,
                    "MTD": mtd_value,
                    "Year": year,
                }
            )

    return records


def parse_workbook(workbook: Workbook) -> tuple[pd.DataFrame, list[str], list[str]]:
    """
    Parse every worksheet in the workbook into a single long-format DataFrame.

    Args:
        workbook: An openpyxl Workbook object (already loaded from bytes).

    Returns:
        A tuple of:
            long_dataframe: pd.DataFrame with columns Sheet, Category,
                Subcategory, Metric, Month, Value, Unit, Target, YTD, MTD, Year.
            available_months: Chronologically ordered list of unique Month labels.
            available_categories: List of unique Category values, in the order
                first encountered in the workbook.
    """
    all_records: list[dict[str, Any]] = []
    month_order: dict[str, tuple[int, int]] = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        if ws.max_row is None or ws.max_row < 1 or ws.max_column is None:
            logger.info("Skipping empty sheet: %s", sheet_name)
            continue

        layout = _detect_layout(ws)
        if layout is None:
            logger.info("Skipping sheet with no detectable KPI table: %s", sheet_name)
            continue

        sheet_records = _parse_sheet(ws, layout)
        if not sheet_records:
            logger.info("Skipping sheet with no data rows: %s", sheet_name)
            continue

        for col in layout.month_cols:
            header_value = ws.cell(row=layout.header_row, column=col).value
            month_abbr, month_num = _month_label_and_num(header_value)
            year = layout.year_map.get(col)
            label = f"{month_abbr}-{year}" if year is not None else month_abbr
            if label not in month_order:
                sort_key = (year if year is not None else 0, month_num)
                month_order[label] = sort_key

        all_records.extend(sheet_records)
        logger.info("Parsed %d records from sheet: %s", len(sheet_records), sheet_name)

    long_dataframe = pd.DataFrame(all_records, columns=LONG_DF_COLUMNS)

    available_months = sorted(month_order.keys(), key=lambda label: month_order[label])

    available_categories = list(
        dict.fromkeys(
            str(cat) for cat in long_dataframe["Category"].tolist() if cat is not None
        )
    )

    return long_dataframe, available_months, available_categories
