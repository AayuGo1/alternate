
"""
excel_loader.py
Handles downloading the Excel workbook from GitHub, caching it via
Streamlit, and converting raw bytes into an openpyxl Workbook object.

This module contains NO business logic, NO parsing of sheet contents,
NO KPI calculations, and NO dataframe creation.
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Final

import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from requests.exceptions import ConnectionError as RequestsConnectionError
from requests.exceptions import Timeout as RequestsTimeout

import config

logger: Final[logging.Logger] = logging.getLogger(__name__)
logger.setLevel(getattr(logging, config.LOG_LEVEL, logging.INFO))

_TIMEOUT_SECONDS: Final[int] = 30
_MAX_RETRIES: Final[int] = 3


class ExcelDownloadError(Exception):
    """Raised when the Excel workbook cannot be downloaded from GitHub."""


class ExcelCorruptedError(Exception):
    """Raised when the downloaded bytes cannot be parsed as a workbook."""


def _validate_url(url: str) -> None:
    """Raise ExcelDownloadError if the configured URL is missing or malformed."""
    if not url or not url.strip():
        raise ExcelDownloadError("RAW_GITHUB_URL in config.py is empty or invalid.")
    if not (url.startswith("http://") or url.startswith("https://")):
        raise ExcelDownloadError(f"RAW_GITHUB_URL is not a valid URL: {url}")


@st.cache_data(ttl=config.RAW_CACHE_TTL_SECONDS, show_spinner=False)
def download_excel() -> bytes:
    """
    Download the Excel workbook from config.RAW_GITHUB_URL.

    Retries up to _MAX_RETRIES times on network failure or timeout.
    Result is cached by Streamlit for config.RAW_CACHE_TTL_SECONDS.

    Returns:
        bytes: Raw content of the downloaded .xlsx file.

    Raises:
        ExcelDownloadError: If the URL is invalid, the file is not found,
            the request times out, or all retry attempts fail.
    """
    url = config.RAW_GITHUB_URL
    _validate_url(url)

    last_error: Exception | None = None

    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            logger.info(
                "Downloading Excel workbook (attempt %d/%d) from %s",
                attempt,
                _MAX_RETRIES,
                url,
            )
          with requests.Session() as session:
            response = session.get(url, timeout=_TIMEOUT_SECONDS)

            if response.status_code == 404:
                raise ExcelDownloadError(
                    f"Workbook not found at {url} (HTTP 404). "
                    f"Verify GITHUB_OWNER, GITHUB_REPO, GITHUB_BRANCH and "
                    f"EXCEL_FILENAME in config.py."
                )

            response.raise_for_status()
            content_type = response.headers.get("Content-Type", "")

if "spreadsheet" not in content_type and "excel" not in content_type:
    logger.warning(
        "Unexpected content type received: %s",
        content_type,
    )
          

            if not response.content:
                raise ExcelDownloadError("Downloaded workbook is empty.")

            logger.info("Successfully downloaded workbook (%d bytes).", len(response.content))
            return response.content

        except ExcelDownloadError:
            raise
        except RequestsTimeout as exc:
            last_error = exc
            logger.warning("Timeout on attempt %d/%d: %s", attempt, _MAX_RETRIES, exc)
        except RequestsConnectionError as exc:
            last_error = exc
            logger.warning("Network error on attempt %d/%d: %s", attempt, _MAX_RETRIES, exc)
        except requests.exceptions.RequestException as exc:
            last_error = exc
            logger.warning("Request failed on attempt %d/%d: %s", attempt, _MAX_RETRIES, exc)

    raise ExcelDownloadError(
        f"Failed to download workbook from {url} after {_MAX_RETRIES} attempts: {last_error}"
    )


def load_workbook_from_bytes(file_bytes: bytes) -> Workbook:
    """
    Convert raw Excel file bytes into an openpyxl Workbook object.

    Args:
        file_bytes: Raw bytes of an .xlsx file.

    Returns:
        Workbook: Parsed openpyxl workbook object.

    Raises:
        ExcelCorruptedError: If the bytes cannot be parsed as a valid
            Excel workbook.
    """
    if not file_bytes:
        raise ExcelCorruptedError("Cannot load workbook: file bytes are empty.")

    try:
        workbook = load_workbook(
    filename=BytesIO(file_bytes),
    data_only=True,
    read_only=True
)
        logger.info("Workbook loaded successfully with %d sheet(s).", len(workbook.sheetnames))
        return workbook
    except Exception as exc:
        logger.error("Failed to parse workbook bytes: %s", exc)
        raise ExcelCorruptedError(f"Workbook appears to be corrupted or invalid: {exc}") from exc


def refresh_cache() -> None:
    """Clear the Streamlit cache so the next call re-downloads the workbook."""
    logger.info("Clearing Excel download cache.")
    download_excel.clear()

def get_workbook() -> Workbook:
    """
    Convenience helper.

    Downloads the workbook and immediately loads it.

    Returns
    -------
    Workbook
    """
    return load_workbook_from_bytes(download_excel())
